# =============================================================================
# modules/excel_writer.py
# =============================================================================

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
from datetime import datetime

KPMG_BLU = "00338D"
BIANCO = "FFFFFF"
NERO = "000000"
ARIAL = "Arial"
FS = 8
FS_Titoli = 14

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border():
    s = Side(style="thin", color=KPMG_BLU)
    return Border(left=s, right=s, top=s, bottom=s)

def _border_analisi(c_idx, start_col, end_col, is_last_row=False):
    s_blu  = Side(style="thin", color=KPMG_BLU)
    s_none = Side(style=None)
    return Border(
        left = s_blu if c_idx == start_col else s_none,
        right = s_blu if c_idx == end_col else s_none,
        top = s_none,
        bottom = s_blu if is_last_row else s_none,
    )

def _border_header(c_idx, start_col, end_col):
    s_blu  = Side(style="thin", color=KPMG_BLU)
    s_none = Side(style=None)
    return Border(
        left = s_blu if c_idx == start_col else s_none,
        right = s_blu if c_idx == end_col   else s_none,
        top = s_blu,
        bottom = s_blu,
    )

def _num_fmt(divisore):
    return '#,##0.0' if divisore > 1 else '#,##0.00'

def _div(val, divisore):
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        return val / divisore
    return val

def _is_perc_col(col_name):
    n = str(col_name).lower()
    return "peso" in n or "var %" in n or "diff %" in n

def _is_nodiv_col(col_name):
    """Colonne che non vanno mai divise per il divisore (duration, convexity, %)."""
    n = str(col_name).lower()
    return any(k in n for k in [
        "dur.",
        "duration",
        "conv.",
        "convexity",
        "(%)",
        "peso",
        "var %",
        "diff %",
        "rendimento",
    ])

def _is_pl_col(col_name):
    n = str(col_name).lower()
    return any(k in n for k in [
        "realizzo", "valutazione", "variazione", "var %",
        "differenza", "diff %", "pl totale",
        "p/l",
        "importo",
        "δp",
        "oci",
        "ecl",
    ])

# Foglio Analisi
def _scrivi_foglio_analisi(ws, df: pd.DataFrame,
                            nome_analisi: str,
                            unita_label: str,
                            divisore: float):
    ws.sheet_view.showGridLines = False
    n_cols = len(df.columns)
    start_col = 2
    end_col = start_col + n_cols - 1
    n_rows = len(df)

    ws.cell(row=2, column=2, value="kpmg").font = Font(
        name="KPMG Logo", size=FS_Titoli, color=KPMG_BLU)
    ws.cell(row=3, column=2, value=nome_analisi).font = Font(
        name="KPMG Bold", size=FS_Titoli, color=KPMG_BLU)

    if end_col > start_col:
        ws.merge_cells(start_row=5, start_column=start_col,
                       end_row=5, end_column=end_col)
    cell_tit = ws.cell(row=5, column=start_col,
                       value=f"{nome_analisi}\n{unita_label}")
    cell_tit.font = Font(name=ARIAL, size=FS, bold=True, color=BIANCO)
    cell_tit.fill = _fill(KPMG_BLU)
    cell_tit.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[5].height = 19.5

    for c_idx, col_name in enumerate(df.columns, start_col):
        cell = ws.cell(row=6, column=c_idx, value=col_name)
        cell.font = Font(name=ARIAL, size=FS, bold=True, color=NERO)
        cell.fill = PatternFill(fill_type=None)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border_header(c_idx, start_col, end_col)
        ws.column_dimensions[get_column_letter(c_idx)].width = max(12, len(str(col_name)) + 2)
    ws.row_dimensions[6].height = 12

    num_fmt = _num_fmt(divisore)
    for r_idx, row in enumerate(df.itertuples(index=False), 7):
        is_tot = str(row[0]).strip().lower() == "totale"
        is_last_row = (r_idx == 6 + n_rows)
        ws.row_dimensions[r_idx].height = 12

        for c_idx, val in enumerate(row, start_col):
            col_name = df.columns[c_idx - start_col]
            is_perc = _is_perc_col(col_name)
            display_val = val if (is_perc or _is_nodiv_col(col_name)) else _div(val, divisore)

            cell = ws.cell(row=r_idx, column=c_idx, value=display_val)
            cell.font = Font(name=ARIAL, size=FS, bold=is_tot, color=NERO)
            cell.border = _border_analisi(c_idx, start_col, end_col, is_last_row)

            if isinstance(display_val, str) and is_perc:
                # Casi speciali _var_pct: "+100%", "-100%", ">100%", "<-100%"
                cell.alignment = Alignment(horizontal="right")
                if display_val.startswith("+") or display_val.startswith(">"):
                    cell.fill = _fill("C6EFCE")
                elif display_val.startswith("-") or display_val.startswith("<"):
                    cell.fill = _fill("FFC7CE")
            elif isinstance(display_val, (int, float)) and not isinstance(display_val, bool):
                cell.number_format = '0.00"%"' if is_perc else num_fmt
                cell.alignment     = Alignment(horizontal="right")
                if _is_pl_col(col_name):
                    if display_val > 0:
                        cell.fill = _fill("C6EFCE")
                    elif display_val < 0:
                        cell.fill = _fill("FFC7CE")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.column_dimensions["A"].width = 2
    ws.freeze_panes = ws.cell(row=7, column=start_col)

# Generazione Workbook
def _scrivi_foglio_raw(ws, df: pd.DataFrame, titolo: str):
    """Scrive un foglio raw con header blu e dati senza divisori."""
    ws.sheet_view.showGridLines = False
    n_cols = len(df.columns)

    for c_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=c_idx, value=str(col_name))
        cell.font = Font(name=ARIAL, size=FS, bold=True, color=BIANCO)
        cell.fill = _fill(KPMG_BLU)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = _border_header(c_idx, 1, n_cols)
        ws.column_dimensions[get_column_letter(c_idx)].width = 16
    ws.row_dimensions[1].height = 14
    ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}1"
    ws.freeze_panes = "A2"

    for r_idx, row in enumerate(df.itertuples(index=False), 2):
        is_last = (r_idx == 1 + len(df))
        ws.row_dimensions[r_idx].height = 12
        for c_idx, val in enumerate(row, 1):
            import math
            if val is pd.NaT or (isinstance(val, float) and math.isnan(val)):
                val = None
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = Font(name=ARIAL, size=FS)
            cell.border = _border_analisi(c_idx, 1, n_cols, is_last)


def genera_excel(dati: dict,
                 nome_portafoglio: str = "Portafoglio",
                 data_report: str = None,
                 unita: str = "€") -> Workbook:

    wb = Workbook()
    data_report = data_report or datetime.today().strftime("%d/%m/%Y")
    divisori = {"€": 1, "€ migliaia": 1_000, "€ milioni": 1_000_000}
    divisore = divisori.get(unita, 1)
    unita_label = f"({unita})"

    # Summary
    ws_r = wb.active
    ws_r.title = "Summary"
    ws_r.sheet_view.showGridLines = False

    ws_r.cell(row=2, column=2, value="kpmg").font = Font(
        name="KPMG Logo", size=FS_Titoli, color=KPMG_BLU)
    ws_r.cell(row=3, column=2, value=nome_portafoglio).font = Font(
        name="KPMG Bold", size=FS_Titoli, color=KPMG_BLU)

    ws_r.merge_cells("B5:E5")
    c = ws_r.cell(row=5, column=2,
                  value=f"Riepilogo Portafoglio\n{data_report} — {unita_label}")
    c.font = Font(name=ARIAL, size=FS, bold=True, color=BIANCO)
    c.fill = _fill(KPMG_BLU)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws_r.row_dimensions[5].height = 19.5

    for i, h in enumerate(["Indicatore", "N", "N-1", "Variazione"], 2):
        cell = ws_r.cell(row=6, column=i, value=h)
        cell.font = Font(name=ARIAL, size=FS, bold=True, color=NERO)
        cell.border = _border()
    ws_r.row_dimensions[6].height = 12

    if "kpi" in dati and dati["kpi"]:
        kpi = dati["kpi"]
        num_fmt = _num_fmt(divisore)
        def _var(n, n1):
            if n is not None and n1 is not None:
                return round(n - n1, 2)
            return None

        rows_kpi = [
            ("NAV Totale", kpi.get("nav"), kpi.get("nav_prev"), kpi.get("var_nav")),
            ("N° Titoli", kpi.get("n_titoli"), kpi.get("n_titoli_prev"), _var(kpi.get("n_titoli"), kpi.get("n_titoli_prev"))),
            ("P&L Totale", kpi.get("pl_totale"), kpi.get("pl_totale_prev"), _var(kpi.get("pl_totale"), kpi.get("pl_totale_prev"))),
            ("Proventi", kpi.get("proventi"), kpi.get("proventi_prev"), _var(kpi.get("proventi"), kpi.get("proventi_prev"))),
            ("Rendimento %", kpi.get("rendimento_%"), kpi.get("rendimento_%_prev"), _var(kpi.get("rendimento_%"), kpi.get("rendimento_%_prev"))),
        ]
        for i, (label, n, n1, var) in enumerate(rows_kpi, 7):
            ws_r.row_dimensions[i].height = 12
            ws_r.cell(row=i, column=2, value=label).font = Font(name=ARIAL, size=FS)
            ws_r.cell(row=i, column=2).border = _border()
            for col_idx, val in enumerate([n, n1, var], 3):
                display = None
                if val is not None:
                    display = val if label in ("N° Titoli", "Rendimento %") else _div(val, divisore)
                cell = ws_r.cell(row=i, column=col_idx, value=display)
                cell.font = Font(name=ARIAL, size=FS)
                cell.border = _border()
                cell.alignment = Alignment(horizontal="right")
                if label == "N° Titoli":
                    cell.number_format = '0'
                elif label == "Rendimento %":
                    cell.number_format = '0.00"%"'
                elif label in ("NAV Totale", "P&L Totale", "Proventi"):
                    cell.number_format = num_fmt
                else:
                    cell.number_format = num_fmt

    for col, w in [("A",2),("B",26),("C",16),("D",16),("E",16)]:
        ws_r.column_dimensions[col].width = w

    # Sheets
    config_fogli = [
        ("patrimoniale_asset_class", "Asset Class", "Composizione per Asset Class"),
        ("patrimoniale_fv_level", "FV Level", "Asset Class per Fair Value Level"),
        ("oci_per_asset_class", "Riserva OCI", "OCI per Asset Class (IFRS9)"),
        ("composizione_valuation_class", "Valuation Class", "Composizione per Valuation Class"),
        ("rating_governativi", "Rating Gov", "Rating – Titoli Governativi"),
        ("rating_non_governativi", "Rating NonGov", "Rating – Titoli Non Governativi"),
        ("geografia_governativi", "Geografia Gov", "Distribuzione Geografica Governativi"),
        ("top_holdings", "Top10 Holdings", "Top 10 Holdings per Book Value"),
        ("top_operazioni", "Top20_Operazioni", "Top 20 Operazioni di Periodo"),
        ("esposizione_valutaria", "Valute", "Esposizione Valutaria"),
        ("esposizione_settoriale", "Settori", "Esposizione Settoriale"),
        ("confronto_bv_fv", "BV vs FV", "Book Value vs Fair Value per Asset Class"),
        ("scadenze_bucket", "Scadenze", "Distribuzione per Bucket di Scadenza"),
        ("duration_ponderata", "Duration", "Duration Ponderata per Asset Class"),
        ("sensitivity_tassi", "Sensitivity", "Stress Test Tassi – Approssimazione di Taylor"),
        ("economica_completa", "Economica", "Analisi Economica per Asset Class"),
        ("effetti_inventory", "Effetti Mkt e Nom", "Effetto Nominale e Mercato (Inventory N vs N-1)"),
        ("effetti_tx_top20", "Top20 Effetti", "Top 20 Operazioni per Effetto Totale"),
        ("effetti_det", "Effetti Op Dettaglio","Effetti per Operazione (Nominale + Prezzo)"),
        ("effetti_rie", "Effetti Op Riepilogo","Riepilogo per ISIN (Nominale + Prezzo + Mercato)"),
    ]

    for chiave, nome_foglio, nome_analisi in config_fogli:
        if chiave not in dati or dati[chiave] is None:
            continue
        df = dati[chiave]
        if isinstance(df, pd.DataFrame) and df.empty:
            continue
        ws = wb.create_sheet(title=nome_foglio)
        _scrivi_foglio_analisi(ws, df,
                               nome_analisi=nome_analisi,
                               unita_label=unita_label,
                               divisore=divisore)

    # Dettaglio Completo
    if "dettaglio" in dati and dati["dettaglio"] is not None:
        ws_det = wb.create_sheet(title="99_Dettaglio")
        ws_det.sheet_view.showGridLines = False

        ETICHETTE = {
            "isin": "ISIN",
            "descrizione": "Descrizione",
            "asset_class": "Asset Class",
            "quantita": "Quantità",
            "prezzo_carico": "Prezzo Carico",
            "book_value": "Book Value N",
            "book_value_prev": "Book Value N-1",
            "fair_value": "Fair Value N",
            "fair_value_prev": "Fair Value N-1",
            "fair_value_level": "Fair Value Level",
            "tipo_emittente": "Tipo Emittente",
            "rating": "Rating",
            "paese": "Paese",
            "valuta": "Valuta",
            "settore": "Settore",
            "cedola": "Interessi N",
            "cedola_prev": "Interessi N-1",
            "dividendi": "Dividendi N",
            "dividendi_prev": "Dividendi N-1",
            "pl_realizzo": "PL Realizzo N",
            "pl_realizzo_prev": "PL Realizzo N-1",
            "pl_valutazione": "PL Valutazione N",
            "pl_valutazione_prev": "PL Valutazione N-1",
            "pl_totale_db": "PL Totale N",
            "pl_totale_db_prev": "PL Totale N-1",
            "oci_lc": "OCI LC N",
            "oci_lc_prev": "OCI LC N-1",
            "ecl_lc": "ECL LC N",
            "ecl_lc_prev": "ECL LC N-1",
            "modified_duration": "Modified Duration",
            "modified_duration_prev": "Modified Duration N-1",
            "convexity": "Convexity",
            "quantita_prev": "Quantità N-1",
            "fair_value_level_prev": "Fair Value Level N-1",
            "data_acquisto": "Data Acquisto",
            "scadenza": "Scadenza",
            "valuation_area": "Valuation Area",
            "company_name": "Società",
            "portfolio_name": "Portafoglio",
            "valuation_class": "Valuation Class",
            "bond_classification": "Bond Classification",
            "security_account_group": "Security Account Group",
        }
        df_det = dati["dettaglio"].rename(columns=ETICHETTE)
        n_det = len(df_det.columns)
        n_det_rows = len(df_det)

        for c_idx, col_name in enumerate(df_det.columns, 1):
            cell = ws_det.cell(row=1, column=c_idx, value=col_name)
            cell.font = Font(name=ARIAL, size=FS, bold=True)
            cell.border = _border_header(c_idx, 1, n_det)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            ws_det.column_dimensions[get_column_letter(c_idx)].width = 16
        ws_det.row_dimensions[1].height = 12
        ws_det.auto_filter.ref = f"A1:{get_column_letter(n_det)}1"
        ws_det.freeze_panes = "A2"

        for r_idx, row in enumerate(df_det.itertuples(index=False), 2):
            is_last = (r_idx == 1 + n_det_rows)
            ws_det.row_dimensions[r_idx].height = 12
            for c_idx, val in enumerate(row, 1):
                cell = ws_det.cell(row=r_idx, column=c_idx, value=val)
                cell.font = Font(name=ARIAL, size=FS)
                cell.border = _border_analisi(c_idx, 1, n_det, is_last)

    # Dati di Input
    raw_config = [
        ("raw_inventory","Inventory","Inventory N + N-1"),
        ("raw_income","Income","Income N + N-1"),
        ("raw_posizioni","Posizioni","Posizioni"),
        ("transaction_report", "TransactionReport", "Transaction Report"),
    ]
    # Check portafoglio effetti (dict -> foglio dedicato)
    if "effetti_check" in dati and dati["effetti_check"]:
        ws_chk = wb.create_sheet(title="EffettiOp_Check")
        ws_chk.sheet_view.showGridLines = False
        chk = dati["effetti_check"]
        ws_chk.cell(row=2, column=2, value="kpmg").font = Font(name="KPMG Logo", size=FS_Titoli, color=KPMG_BLU)
        ws_chk.cell(row=3, column=2, value="Check Riconciliazione Effetti").font = Font(name="KPMG Bold", size=FS_Titoli, color=KPMG_BLU)
        for r_idx, (k, v) in enumerate(chk.items(), 5):
            ws_chk.cell(row=r_idx, column=2, value=k).font = Font(name=ARIAL, size=FS, bold=True)
            cell = ws_chk.cell(row=r_idx, column=3, value=v)
            cell.font = Font(name=ARIAL, size=FS)
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal="right")
            if k == "Check Portafoglio":
                cell.fill = _fill("C6EFCE") if abs(v or 0) < 1000 else _fill("FFC7CE")
        for col, w in [("A",2),("B",30),("C",18)]:
            ws_chk.column_dimensions[col].width = w

    for chiave, nome_foglio, titolo in raw_config:
        if chiave not in dati or dati[chiave] is None:
            continue
        df_raw = dati[chiave]
        if isinstance(df_raw, pd.DataFrame) and df_raw.empty:
            continue
        ws_raw = wb.create_sheet(title=nome_foglio)
        _scrivi_foglio_raw(ws_raw, df_raw, titolo)

    return wb


def salva_excel(wb: Workbook, path: str):
    wb.save(path)
    print(f"[OK] Excel salvato: {path}")
